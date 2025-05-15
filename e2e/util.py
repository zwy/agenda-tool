import pandas as pd
import openpyxl
import json
import os
import numpy as np
from io import BytesIO

def convert_value(x):
    """
    处理数据值，将NaN转换为空字符串，其他值转为字符串并去除首尾空格
    
    Args:
        x: 要处理的值
    
    Returns:
        处理后的字符串值
    """
    if pd.isna(x) or x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    return str(x).strip()

def process_excel(excel_path):
    """
    处理Excel文件，解决合并单元格问题，并将数据转换为字符串
    
    Args:
        excel_path: Excel文件的绝对路径
    
    Returns:
        数据列表，每一行为一个字典
    """
    # 首先使用openpyxl加载工作簿来处理合并单元格
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active  # 假设我们只处理第一个工作表
    
    # 获取所有合并单元格的信息并存储
    merged_cell_info = []
    for merged_range in sheet.merged_cells.ranges:
        # 获取左上角单元格的值
        top_left_cell_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
        merged_cell_info.append((merged_range, top_left_cell_value))
    
    # 解除所有单元格的合并
    for merged_range, _ in merged_cell_info:
        sheet.unmerge_cells(str(merged_range))
    
    # 为之前合并的单元格填充相同的值
    for merged_range, value in merged_cell_info:
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                sheet.cell(row=row, column=col).value = value
    
    # 创建一个BytesIO对象，用于在内存中保存处理后的Excel
    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    
    # 使用pandas读取处理后的Excel
    df = pd.read_excel(virtual_workbook)
    
    # 处理列名，去除或重命名 "Unnamed:" 列
    for col in df.columns:
        if isinstance(col, str) and col.startswith('Unnamed:'):
            # 可以选择删除这些列或给它们一个更有意义的名称
            # 这里我们选择删除
            df = df.drop(columns=[col])
    
    # 将所有数据转换为字符串，并处理NaN
    df = df.applymap(convert_value)
    
    # 将DataFrame转换为字典列表
    data_list = df.to_dict(orient='records')
    
    return data_list

def save_to_jsonl(data, output_file):
    """
    将数据保存为JSONL文件
    
    Args:
        data: 要保存的数据列表
        output_file: 输出文件路径
    """
    # 确保目录存在
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # 写入JSONL文件
    with open(output_file, 'w', encoding='utf-8') as f:
        for item in data:
            f.write(json.dumps(item, ensure_ascii=False) + '\n')
    
    print(f"数据已成功保存到 {output_file}")
    print(f"数据总行数: {len(data)}")
    
    # 打印样例数据，以便确认处理结果
    if data:
        print("\n样例数据 (第一行):")
        print(json.dumps(data[0], ensure_ascii=False, indent=2))

def filter_fields(data, fields_to_keep):
    """
    裁剪数据，只保留指定的字段
    
    Args:
        data: 原始数据列表，每项为一个字典
        fields_to_keep: 要保留的字段名列表
    
    Returns:
        裁剪后的数据列表
    """
    filtered_data = []
    
    for item in data:
        filtered_item = {}
        for field in fields_to_keep:
            if field in item:
                filtered_item[field] = item[field]
        filtered_data.append(filtered_item)
    
    return filtered_data

def rename_fields(data, field_map):
    """
    重命名数据中的字段
    
    Args:
        data: 原始数据列表，每项为一个字典
        field_map: 字段映射字典，键为原字段名，值为新字段名
    
    Returns:
        字段重命名后的数据列表
    """
    renamed_data = []
    
    for item in data:
        renamed_item = {}
        for key, value in item.items():
            # 如果当前字段在映射字典中，使用新名称，否则保持原名
            new_key = field_map.get(key, key)
            renamed_item[new_key] = value
        renamed_data.append(renamed_item)
    
    return renamed_data

def save_to_excel(data, output_file):
    """
    将数据保存为Excel文件
    
    Args:
        data: 要保存的数据列表，每项为一个字典
        output_file: 输出Excel文件路径
    """
    # 确保目录存在
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # 将数据转换为DataFrame
    df = pd.DataFrame(data)
    
    # 保存为Excel文件
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='数据')
        
        # 获取工作簿和工作表对象，以便进行格式调整
        workbook = writer.book
        worksheet = writer.sheets['数据']
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # 稍微增加宽度以便更好地显示
            adjusted_width = (max_length + 2)
            # 限制最大宽度，避免过宽
            adjusted_width = min(adjusted_width, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"数据已成功保存到Excel文件: {output_file}")
    print(f"数据总行数: {len(data)}")
    print(f"数据总列数: {len(data[0]) if data else 0}")
